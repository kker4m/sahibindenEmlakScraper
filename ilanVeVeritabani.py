
import sqlite3
from bs4 import BeautifulSoup
from selenium import webdriver
import os
import time
import openpyxl


conn=sqlite3.connect("main.db")
cursor = conn.cursor()

sql_command='''
        CREATE TABLE IF NOT EXISTS ilanlar(
        ilanadı TEXT,ozellik1 TEXT,ozellik2 TEXT,ozellik3 TEXT,ozellik4 TEXT,ozellik5 TEXT,ozellik6 TEXT,ozellik7 TEXT,ozellik8 TEXT,ozellik9 TEXT,ozellik10 TEXT,ozellik11 TEXT,ozellik12 TEXT,ozellik13 TEXT,ozellik14 TEXT,ozellik15 TEXT,ozellik16 TEXT,ozellik17 TEXT,ozellik18 TEXT,ozellik19 TEXT,ozellik20 TEXT,ozellik21 TEXT,ozellik22 TEXT,ozellik23 TEXT,ozellik24 TEXT,ozellik25 TEXT, ozellik26 TEXT,ozellik27 TEXT
)'''

cursor.execute(sql_command)

def trtoen(text):
    if text=="İstanbul":
        return "istanbul"
    if text=="Köşk & Konak":
        return "kosk-konak"
    onceki_karakter = ["ş", "ç", "ö", "ğ", "ü", "ı"," "]
    sonraki_karakter = ["s", "c", "o", "g", "u", "i","-"]
    if text[0]=="i":
        text="i"+text[1:]
    text=str(text)
    text=text.lower()
    for i in range(7):
        text=text.replace(onceki_karakter[i],sonraki_karakter[i])
    return text


def sayfasayisi():
    r = chrome.page_source
    soup = BeautifulSoup(r, "lxml")
    sayfasayisi=soup.find("p",attrs={"class":"mbdef"})
    sayfa=sayfasayisi.text
    sayfa=sayfa.split()
    sayfa=sayfa[1]
    return str(sayfa)#
def resimadet(soup):
    resimadet=soup.find("span",attrs={"class":"images-count"})
    resimadet=resimadet.text
    resimadet=resimadet[3:5]
    resimadet=resimadet.replace("/","")
    return str(resimadet)+" Adet resim var"



chromeOptions= webdriver.ChromeOptions()
#chromeOptions.add_argument("--headless")
chromeOptions.add_argument("--no-sandbox")
chromeOptions.add_argument("start-maximized")
chromeOptions.add_argument("disable-infobars")
chromeOptions.add_argument("--disable-extensions")
current_path=os.path.dirname(os.path.abspath(__file__))+"\chromedriver.exe"
current_path=current_path.replace("\\","\\\\")
chrome=webdriver.Chrome(executable_path=str(current_path),options=chromeOptions)



wb_obj=openpyxl.load_workbook("main.xlsm")
sheet_obj=wb_obj.active

ilantipi_excel=sheet_obj.cell(row=1,column=1)
konuttipi_excel=sheet_obj.cell(row=1,column=2)
il_excel=sheet_obj.cell(row=1,column=3)

ilantipi_excel=ilantipi_excel.value
ilantipi_excel=trtoen(ilantipi_excel)

konuttipi_excel=konuttipi_excel.value
konuttipi_excel=trtoen(konuttipi_excel)

il_excel=il_excel.value
il_excel=trtoen(il_excel)

link="https://www.sahibinden.com/"+str(ilantipi_excel)+"-"+str(konuttipi_excel)+"/"+str(il_excel)







chrome.get(link)
try:
    pagecount=sayfasayisi()
except:
    pagecount=1
r=chrome.page_source
soup = BeautifulSoup(r,"lxml")
lastindex=0


for page in range(int(pagecount)):
    chrome.get(link+"?pagingOffset="+str(lastindex))
    ilanbody=soup.find("table",attrs={"id":"searchResultsTable"})
    ilanlar=ilanbody.find_all("tr")

    for ilan in ilanlar:

        #İlanın bulunması

        ozelliklist=[]

        ilanchild=ilan.find("td",attrs={"class":"searchResultsLargeThumbnail"})
        try:
            urunAdi=ilanchild.a.get("title")
            urunlink=ilanchild.a.get("href")
            ozelliklist.append(urunAdi)
        except: continue
        time.sleep(5)
        chrome.get(url=("https://www.sahibinden.com"+urunlink))
        urunlinkkaynak=chrome.page_source
        urunsoup=BeautifulSoup(urunlinkkaynak,"lxml")
        ilanozellik=urunsoup.find("ul",attrs={"class":"classifiedInfoList"})

        #Satıcı ve firma özellikleri
        try:
            user_all_info=urunsoup.find("div",attrs={"class":"user-info-module"})
            ahrefisim=user_all_info.find("div",attrs={"class","user-info-store-card"})
            isim_store = ahrefisim.a.get("title")
            isim_kullanici_1 = user_all_info.find("div", attrs={"class", "user-info-agent"})
            isim_kullanici_2 = isim_kullanici_1.find("h3")
            isim_kullanici=isim_kullanici_2.text
            telefonnumara = user_all_info.find("div", attrs={"class": "user-info-phones"})
            telefnnumaratext = telefonnumara.text
            isim_store=isim_store.strip()
            isim_kullanici=isim_kullanici.strip()
            telefnnumaratext=telefnnumaratext.strip()
            telefnnumaratext=telefnnumaratext.replace("\n","")
            ozelliklist.append(isim_store)
            ozelliklist.append(isim_kullanici)
            ozelliklist.append(telefnnumaratext)
        except:pass

        # İlan Özellikleri


        ozellikler=ilanozellik.find_all("li")
        ozelliklist.append(resimadet(urunsoup))
        for ozellik in ozellikler:
            a=ozellik.text
            a=a.replace("\n","")
            a=a.replace(" ","")
            a=a.replace("\xa0"," = ")
            a=a.replace("\t","")
            ozelliklist.append(a)
        while len(ozelliklist)<28:
            ozelliklist+=" "


        cursor.execute("INSERT INTO ilanlar VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)",ozelliklist)
        conn.commit()
    lastindex+=20

conn.close()









